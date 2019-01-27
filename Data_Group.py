import scipy.sparse
import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
import xlsxwriter
import os
from tqdm import tqdm
import time

class RadarImage(object):
    def __init__(self, file_path=None):
        self.file_path= file_path
        os.chdir(self.file_path)
        self.pred_file_name = os.listdir(self.file_path)
        self.mat=None

    def read_mat(self, npc_file):
        mat= scipy.sparse.load_npz(npc_file).todense()
        return mat

    def extract_ref_time(self, file_name):
        _, base_time, lead_time, pred_time = file_name.split('_')
        pred_time = pred_time[:-4]
        return base_time, lead_time, pred_time

    def retrieve_leadtime_lists(self, base_time):
        file_names = []
        for file in self.pred_file_name:
            if file.split("_")[1] == base_time:
                file_names.append(file)
        return file_names

    def moving_window(self, matrix,  patch_size=50):
        # assume matrix is regular
        x,y = matrix.shape
        for i in range(int(x/patch_size)):
            for j in range(int(y/patch_size)):
                yield matrix[i:i+patch_size, j:j+patch_size],(i*50,j*50)
    
    def find_uniq_file(self):
        base_times = [self.pred_file_name[i].split('_')[1] for i in range(len(self.pred_file_name))]
        self.base_times = set(base_times)
        return None

    def execute(self):
        for i, base_time in tqdm(enumerate(self.base_times)):
            start = time.time()
            print( f'----processing {base_time}----')
            lead_files = self.retrieve_leadtime_lists(base_time)
            for j, lead_file in enumerate(lead_files):
                lead_time = lead_file.split("_")[-1][:-4]
                matrix = self.read_mat(lead_file)
                # print(f'subprocess {lead_time}')
                for sub_mat, tup in self.moving_window(matrix):
                    try:
                        wb = load_workbook(f'../Table/test-{tup}.xlsx')
                        ws = wb.worksheets[0]
                        ws.cell(row=i+1, column=j+1, value=sub_mat.mean())
                        wb.save(f'../Table/test-{tup}.xlsx')
                    except FileNotFoundError:
                        wb = Workbook()
                        ws = wb.active
                        ws.cell(row=i+1, column=j+1, value=sub_mat.mean())
                        wb.save(f'../Table/test-{tup}.xlsx')
                    finally:
                        wb.close()
            end = time.time()
            elapsed = end-start
            print(f"----------one echo done! elapsed time {elapsed/3600.}----------")

    def synthesize_workbook(self):
        files = os.listdir('../Table/')
        for file in files:
            df = pd.read_excel(f"../Table/{file}", header=None)
            df.columns = [str(i)+'min lead' for i in range(2,62,2)]
            df['index'] = self.base_times
            df.set_index('index', inplace=True)
            df.to_excel(f'../Table/{file}')

def main():
    start = time.time()
    RI = RadarImage('E:\\Learning Materials\\Radar-h2i\\DATA')
    RI.find_uniq_file()
    RI.execute()
    RI.synthesize_workbook()
    end = time.time()
    elapsed = end-start
    print(f"------------Done! elapsed time : {elapsed/3600.}---------------")

if __name__ =='__main__':
    main()
